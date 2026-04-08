import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# ---------------- FILE PATH FIX ----------------

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = os.path.join(BASE_DIR, "app", "incident_reports.xlsx")


def create_file():

    wb = Workbook()
    ws = wb.active

    headers = [
        "timestamp",
        "reporter_name",
        "department",
        "equipment",
        "incident_summary",
        "location_or_unit",
        "incident_date",
        "incident_time",
        "severity",
        "measured_parameters",
        "remarks"
    ]

    ws.append(headers)

    wb.save(FILE_NAME)


def save_report_to_excel(data):

    try:

        # Ensure folder exists
        os.makedirs(os.path.dirname(FILE_NAME), exist_ok=True)

        if not os.path.exists(FILE_NAME):
            create_file()

        wb = load_workbook(FILE_NAME)
        ws = wb.active

        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            data.get("reporter_name",""),
            data.get("department",""),
            data.get("equipment",""),
            data.get("incident_summary",""),
            data.get("location_or_unit",""),
            data.get("incident_date",""),
            data.get("incident_time",""),
            data.get("severity",""),
            str(data.get("measured_parameters","")),
            data.get("remarks","")
        ]

        ws.append(row)

        wb.save(FILE_NAME)

    except Exception as e:
        print("Excel write error:", e)